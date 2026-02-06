from math import exp
from openpyxl import load_workbook
import bpy
import os

def parsed_filepath(filepath):
    base_name = os.path.basename(filepath)
    name, ext = os.path.splitext(base_name)
    parsed = name.split('_')
    return parsed, ext

def increment_number(parsed_filepath, layout=False):
    parsed, ext = parsed_filepath
    # gestion vXX
    if len(parsed) > 1 and parsed[-2].startswith("v"):
        new_version =  int("01")
        parsed[-2] = f"v{new_version:02d}"
        if layout:
            parsed[-3] = "An"
        new_name = '_'.join(parsed) + ext

    return new_name

def find_dir_ep(blend_path):
    blend_path = bpy.data.filepath
    dirs = blend_path.split(os.sep)
    for i, d in enumerate(dirs):
        if d.isdigit():
            ep_dir = os.sep.join(dirs[:i+1])
            return ep_dir
        
def find_xlsx(ep_dir):
    ep = os.path.basename(find_dir_ep(ep_dir))
    xlsx_path = os.path.join(ep_dir,f"EP{ep}_Material" ,"DOCS",f'MM_{ep}_VB_01_shotlist.xlsx')
    # print("xlsx_path",xlsx_path)
    return xlsx_path

def match_shot():
    blend_path = bpy.data.filepath
    filename = find_xlsx(find_dir_ep(blend_path))
    workbook = load_workbook(filename=filename, data_only=True)
    
    blend_name = bpy.path.basename(blend_path)
    ep = blend_name.split('_')[1]
    #Check first column to find the good asset_txt = row[j].value and not set it manually
    for j, col in enumerate(workbook[f'MM_{ep}_VB_01'].iter_cols()):
        if col[0].value != 'SceneName' : continue 
        # print(col[0].value )
        shot_list = [row[j].value for i, row in enumerate(workbook[f'MM_{ep}_VB_01'].iter_rows()) if isinstance(row[j].value, (int, float))]          
    return shot_list

def find_new_shot(blend_path):
    #If a next shot exist in this episode, return this next shot
    blend_path = bpy.data.filepath
    dirs = blend_path.split(os.sep)
    if not "Layout" in dirs :
        return None
    for i, d in enumerate(dirs):
        if d.isdigit():
            shot = str("{:04d}".format(int(d)+10))
            new_shot_dir = os.sep.join(dirs[:i] + [shot])
            ep_dir = os.sep.join(dirs[:i+1])
            parsed = os.path.basename(bpy.data.filepath).split("_")
            if len(parsed) > 1 and parsed[-2].startswith("v"):
                parsed[-2] = "V01"
                parsed[3] = shot
            new_name = "_".join(parsed)
            print(os.path.join(new_shot_dir, "Layout"))
    new_shot = int(shot)
    
    shot_list = match_shot()
    if not new_shot in list(shot_list) :
        print("new_shot", new_shot)
        print("match ",shot_list)
        return None

    if not os.path.exists(os.path.join(new_shot_dir, "Layout")):
        os.makedirs(os.path.join(new_shot_dir, "Layout"))
        print(f"dossier {os.path.join(new_shot_dir, 'Layout')} a été crée")

    return os.path.join(new_shot_dir, "Layout" ,new_name)
