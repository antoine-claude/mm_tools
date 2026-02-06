from math import exp
from openpyxl import load_workbook
import bpy
import os

def find_prod_dir(blend_path):
    dirs = blend_path.split(os.sep)
    for i, d in enumerate(dirs):
        if d.isdigit():
            ep_dir = os.sep.join(dirs[:i-1])
            return ep_dir
        

def find_dir_ep(blend_path):
    dirs = blend_path.split(os.sep)
    for i, d in enumerate(dirs):
        if d.isdigit():
            ep_dir = os.sep.join(dirs[:i+1])
            return ep_dir
        

def find_xlsx(blend_path):
    ep_dir = find_dir_ep(blend_path)
    ep = os.path.basename(ep_dir)
    xlsx_path = os.path.join(ep_dir,f"EP{ep}_Material" ,"DOCS",f'MM_{ep}_VB_01_shotlist.xlsx')
    # print("xlsx_path",xlsx_path)
    return xlsx_path


def match_shot():
    blend_path = bpy.data.filepath
    filename = find_xlsx(blend_path)
    workbook = load_workbook(filename=filename, data_only=True)
    
    blend_name = bpy.path.basename(blend_path)
    ep = blend_name.split('_')[1]
    #Check first column to find the good asset_txt = row[j].value and not set it manually
    for j, col in enumerate(workbook[f'MM_{ep}_VB_01'].iter_cols()):
        if col[0].value != 'assetListAnimation' : continue 
        # print(col[0].value )
        for i, row in enumerate(workbook[f'MM_{ep}_VB_01'].iter_rows()):
            num = row[2].value
            asset_txt = row[j].value
            if not isinstance(num, (int, float)):continue
            if blend_name.split('_')[3] == f"{int(num):04d}":
                assets = [p.strip() for p in asset_txt.split('-') if p.strip()]
                # print(f"yes chef {assets} correspond a blendfile {int(num):04d} au row {i+1}")
    assets.append("MM_Camera")
    return assets, num


def find_file(match_shot):
    candidates = []
    assets, num = match_shot
    # print(assets, "num :", num)
    prod_dir = find_prod_dir(bpy.data.filepath)
    base_chars = os.path.join(prod_dir, "Assets", "Characters")
    base_props = os.path.join(prod_dir, "Assets", "Props")
    base_sets = os.path.join(prod_dir, "Assets", "Sets")
    base_cam = os.path.join(prod_dir, "Assets", "Camera")
    base_set_items = os.path.join(prod_dir, "Assets", "Set_Items")
    final_rd_path = "Final\Render"
    for asset in assets:
        if asset.split('_')[1] == 'CHR':
            candidate = os.path.join(base_chars, asset, final_rd_path, f'{asset}.blend')
            # print(candidate)
        elif asset.split('_')[1] == 'PRP':
            candidate = os.path.join(base_props, asset, final_rd_path, f'{asset}.blend')
            # print(os.path.join(base_props, asset, final_rd_path, f'{asset}.blend'))
        elif asset.split('_')[1] == 'ITM':
            candidate = os.path.join(base_set_items, asset, final_rd_path, f'{asset}.blend')
        elif asset.split('_')[1] == 'SET':
            candidate = os.path.join(base_sets, asset, final_rd_path, f'{asset}.blend')
            # print("SET",os.path.join(base_sets, asset, final_rd_path, f'{asset}.blend'))
        elif asset.split('_')[1] == 'Camera':
            candidate = os.path.join(base_cam, f'{asset}.blend')
            # print("SET",os.path.join(base_cam, asset, final_rd_path, f'{asset}.blend'))
        else :
            candidate = None
        if not candidate:continue
        if os.path.exists(candidate):
            candidates.append(candidate)
        
            
    return candidates

def find_root_collections(data_from):
    """
    Retourne les collections sans parent
    (équivalent logique d'un top-level scene collection)
    """
    parents = set()

    for col in data_from.collections:
        for child in col.children:
            parents.add(child)

    root_cols = [
        col for col in data_from.collections
        if col not in parents
    ]

    return root_cols


def link_collection_matching_filename(blend_path):
    """
    Link la collection qui a le même nom que le fichier .blend
    """

    if not os.path.exists(blend_path):
        print(f"[ERROR] .blend introuvable : {blend_path}")
        return None

    expected_name = os.path.splitext(os.path.basename(blend_path))[0]
    is_env = False
    if expected_name.split('_')[1] == 'CHR':
        expected_col = next((c for c in bpy.data.collections if c.name.casefold() == 'chara'.casefold()), None)
        if expected_col is None :
            expected_col = bpy.data.collections.new("Chara")
            bpy.context.scene.collection.children.link(expected_col)
            expected_col.color_tag = 'COLOR_01'
    elif expected_name.split('_')[1] == 'PRP':
        expected_col = next((c for c in bpy.data.collections if c.name.casefold() == 'props'.casefold()), None)
        if expected_col is None :
            expected_col = bpy.data.collections.new("Props")
            bpy.context.scene.collection.children.link(expected_col)
            expected_col.color_tag = 'COLOR_05'
    elif expected_name.split('_')[1] == 'SET':
        expected_col = next((c for c in bpy.data.collections if c.name.casefold() == 'env'.casefold()), None)
        expected_name = os.path.splitext(os.path.basename(blend_path))[0].split("_")[2]
        is_env = True
        if expected_col is None :
            expected_col = bpy.data.collections.new("Env")
            bpy.context.scene.collection.children.link(expected_col)
    elif expected_name.split('_')[1] == "Camera" :
        expected_col = next((c for c in bpy.data.collections if c.name.casefold() == 'cam'.casefold()), None)
        if expected_col is None :
            expected_col = bpy.data.collections.new("Cam")
            expected_col.color_tag = 'COLOR_04'
            bpy.context.scene.collection.children.link(expected_col)

    try:
        with bpy.data.libraries.load(blend_path, link=True, relative=False) as (data_from, data_to):

            if not is_env:
                print("Link d'un asset")

                if expected_name not in data_from.collections:
                    print(f"[WARNING] '{expected_name}' absente de {blend_path}")
                    return None

                # cas simple : on link direct par nom
                data_to.collections = [expected_name]

            else:
                print("Link d'un env")

                # 1) on charge TOUTES les collections
                data_to.collections = list(data_from.collections)

        # 2) post-traitement APRÈS le with
        if is_env:
            loaded_cols = data_to.collections
            child_names = set()

            for col in loaded_cols:
                for child in col.children:
                    child_names.add(child.name)

            root_cols = [col for col in loaded_cols if col.name not in child_names]

            if not root_cols:
                print("[ERROR] Aucune collection racine trouvée")
                return None

            linked_col = root_cols[0]

        else:
            linked_col = data_to.collections[0]
            

        if not is_env :
            override_col = linked_col.override_hierarchy_create(
                scene=bpy.context.scene,
                view_layer=bpy.context.view_layer,
                do_fully_editable=True
        
        )
        
        else :
            override_col = linked_col
        if expected_col is not None:
            if override_col.name not in expected_col.children:
                expected_col.children.link(override_col)
        if override_col in bpy.context.scene.collection.children_recursive:
            bpy.context.scene.collection.children.unlink(override_col)
        print(f"[OK] {expected_name} linkée depuis {blend_path}")
        return override_col

    except Exception as e:
        print(f"[ERROR] Échec du link '{expected_name}'")
        print(f"        {e}")
        return None



# def unlink_collection_not_matching_filename()